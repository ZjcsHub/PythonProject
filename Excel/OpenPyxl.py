# coding=utf-8

from openpyxl import load_workbook
from openpyxl.styles import colors,Font,Fill,NamedStyle
from openpyxl.styles import PatternFill,Alignment,Border,Side
from openpyxl.worksheet.worksheet import Worksheet


def read_file(fileName):
    #加载文件
    wb = load_workbook(fileName)
    #读取sheetname
    print('输出文件所有工作表名 :\n',wb.sheetnames)
    ws = wb['Sheet1']
    # 如果不知道名字时
    sheet_names = wb.sheetnames
    ws2 = wb[sheet_names[0]]
    print(ws is ws2)

    #修改sheetname
    ws.title = 'iOS'
    print("修改sheetname", wb.sheetnames)

    #创建新的sheet
    #创建的新表必须要赋值给一个对象，不然只有名字没有实际的表
    ws3 = wb.create_sheet('iOS',0)
    #什么参数都不写的话，默认插入到最后一个位置且名字为sheet，sheet1，。。。
    ws4 = wb.create_sheet()

    print('创建新的sheet：\n',wb.sheetnames)

    # 删除sheet
    wb.remove(ws3)

    print("删除sheet后的sheet :\n",wb.sheetnames)

    # 修改sheet选项卡背景色，默认为白色，设置RRGGBB模式
    ws.sheet_properties.tabColor = "FFA500"

    # 读取有效区域
    print("最大列数",ws.max_column)
    print("最大行数", ws.max_row)


    # 读取
    c = ws['C1']
    print(c.value)

    # 插入行和列
    ws.insert_rows(1)
    ws.insert_cols(2,4)

    #删除行和列
    ws.delete_cols(6,3)
    ws.delete_rows(3)

def cell_operation(fileName):
    # 加载文件
    wb = load_workbook(fileName)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]

    # 读取
    c = ws['B2']
    c1 = ws.cell(row=1,column=2)
    print(c,c1)
    print(c.value,c1.value)

    # 修改
    ws['B2'] = 'TEst app'
    ws.cell(1,2).value = "2020/01/03"

    print(c.value,c1.value)

    #读取多个单元格
    cell_range = ws['A1':'D5']
    print("A1:D5 :",cell_range)
    #A1:D5 : ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <MergedCell 'Sheet1'.C1>, <MergedCell 'Sheet1'.D1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.D2>), (<Cell 'Sheet1'.A3>, <MergedCell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.D3>), (<Cell 'Sheet1'.A4>, <MergedCell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>, <Cell 'Sheet1'.D4>), (<Cell 'Sheet1'.A5>, <MergedCell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>, <Cell 'Sheet1'.D5>))
    colC = ws['C']
    print("ws[C]:", colC)
    #ws[C]: (<MergedCell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>, <Cell 'Sheet1'.C5>, <Cell 'Sheet1'.C6>, <Cell 'Sheet1'.C7>, <Cell 'Sheet1'.C8>, <Cell 'Sheet1'.C9>, <Cell 'Sheet1'.C10>, <Cell 'Sheet1'.C11>, <Cell 'Sheet1'.C12>, <Cell 'Sheet1'.C13>, <Cell 'Sheet1'.C14>, <Cell 'Sheet1'.C15>, <Cell 'Sheet1'.C16>, <MergedCell 'Sheet1'.C17>, <MergedCell 'Sheet1'.C18>, <MergedCell 'Sheet1'.C19>, <MergedCell 'Sheet1'.C20>, <Cell 'Sheet1'.C21>, <MergedCell 'Sheet1'.C22>, <MergedCell 'Sheet1'.C23>, <Cell 'Sheet1'.C24>, <MergedCell 'Sheet1'.C25>, <Cell 'Sheet1'.C26>)

    col_range = ws['C':'D']
    print("ws['C':'D']:", col_range)
    #ws['C':'D']: ((<MergedCell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>, <Cell 'Sheet1'.C5>, <Cell 'Sheet1'.C6>, <Cell 'Sheet1'.C7>, <Cell 'Sheet1'.C8>, <Cell 'Sheet1'.C9>, <Cell 'Sheet1'.C10>, <Cell 'Sheet1'.C11>, <Cell 'Sheet1'.C12>, <Cell 'Sheet1'.C13>, <Cell 'Sheet1'.C14>, <Cell 'Sheet1'.C15>, <Cell 'Sheet1'.C16>, <MergedCell 'Sheet1'.C17>, <MergedCell 'Sheet1'.C18>, <MergedCell 'Sheet1'.C19>, <MergedCell 'Sheet1'.C20>, <Cell 'Sheet1'.C21>, <MergedCell 'Sheet1'.C22>, <MergedCell 'Sheet1'.C23>, <Cell 'Sheet1'.C24>, <MergedCell 'Sheet1'.C25>, <Cell 'Sheet1'.C26>), (<MergedCell 'Sheet1'.D1>, <Cell 'Sheet1'.D2>, <Cell 'Sheet1'.D3>, <Cell 'Sheet1'.D4>, <Cell 'Sheet1'.D5>, <Cell 'Sheet1'.D6>, <Cell 'Sheet1'.D7>, <Cell 'Sheet1'.D8>, <Cell 'Sheet1'.D9>, <Cell 'Sheet1'.D10>, <Cell 'Sheet1'.D11>, <Cell 'Sheet1'.D12>, <Cell 'Sheet1'.D13>, <Cell 'Sheet1'.D14>, <Cell 'Sheet1'.D15>, <Cell 'Sheet1'.D16>, <Cell 'Sheet1'.D17>, <Cell 'Sheet1'.D18>, <Cell 'Sheet1'.D19>, <Cell 'Sheet1'.D20>, <Cell 'Sheet1'.D21>, <Cell 'Sheet1'.D22>, <Cell 'Sheet1'.D23>, <Cell 'Sheet1'.D24>, <Cell 'Sheet1'.D25>, <Cell 'Sheet1'.D26>))

    row10 = ws[10]
    print("ws[10]:", row10)
    #ws[10]: (<Cell 'Sheet1'.A10>, <MergedCell 'Sheet1'.B10>, <Cell 'Sheet1'.C10>, <Cell 'Sheet1'.D10>, <MergedCell 'Sheet1'.E10>, <MergedCell 'Sheet1'.F10>, <MergedCell 'Sheet1'.G10>, <Cell 'Sheet1'.H10>)

    row_range = ws[5:10]
    print('row_range :',row_range)
    #row_range : ((<Cell 'Sheet1'.A5>, <MergedCell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>, <Cell 'Sheet1'.D5>, <MergedCell 'Sheet1'.E5>, <MergedCell 'Sheet1'.F5>, <MergedCell 'Sheet1'.G5>, <Cell 'Sheet1'.H5>), (<Cell 'Sheet1'.A6>, <MergedCell 'Sheet1'.B6>, <Cell 'Sheet1'.C6>, <Cell 'Sheet1'.D6>, <MergedCell 'Sheet1'.E6>, <MergedCell 'Sheet1'.F6>, <MergedCell 'Sheet1'.G6>, <Cell 'Sheet1'.H6>), (<Cell 'Sheet1'.A7>, <MergedCell 'Sheet1'.B7>, <Cell 'Sheet1'.C7>, <Cell 'Sheet1'.D7>, <MergedCell 'Sheet1'.E7>, <MergedCell 'Sheet1'.F7>, <MergedCell 'Sheet1'.G7>, <Cell 'Sheet1'.H7>), (<Cell 'Sheet1'.A8>, <MergedCell 'Sheet1'.B8>, <Cell 'Sheet1'.C8>, <Cell 'Sheet1'.D8>, <MergedCell 'Sheet1'.E8>, <MergedCell 'Sheet1'.F8>, <MergedCell 'Sheet1'.G8>, <Cell 'Sheet1'.H8>), (<Cell 'Sheet1'.A9>, <Cell 'Sheet1'.B9>, <Cell 'Sheet1'.C9>, <Cell 'Sheet1'.D9>, <MergedCell 'Sheet1'.E9>, <MergedCell 'Sheet1'.F9>, <MergedCell 'Sheet1'.G9>, <Cell 'Sheet1'.H9>), (<Cell 'Sheet1'.A10>, <MergedCell 'Sheet1'.B10>, <Cell 'Sheet1'.C10>, <Cell 'Sheet1'.D10>, <MergedCell 'Sheet1'.E10>, <MergedCell 'Sheet1'.F10>, <MergedCell 'Sheet1'.G10>, <Cell 'Sheet1'.H10>))


    # 按照行列操作
    for row in ws.iter_rows(min_row=1,max_row=3,min_col=1,max_col=3):
        # print(row)
        for cell in row:
            print(cell)



def cell_operation2(fileName):
    # 加载文件
    wb = load_workbook(fileName)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]

    # 合并单元格
    ws.merge_cells('H3:H4')
    ws['H3'] = '合并两个单元格'
    # 或者
    ws.merge_cells(start_row=5,start_column=8,end_row=8,end_column=8)
    ws.cell(2,6).value = "合并3个单元格"

    #取消合并单元格
    # ws.unmerge_cells('H3:H4')
    #或者
    # ws.unmerge_cells(start_row=5,start_column=8,end_row=8,end_column=8)
    wb.save(fileName)

if __name__ == "__main__":
    cell_operation2("/Users/app005synergy/Desktop/ios用户评论汇总0225.xlsx")