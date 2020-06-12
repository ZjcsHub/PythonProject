import xlrd
import os
from pip._vendor.distlib.compat import raw_input
from openpyxl import load_workbook,Workbook

wb = Workbook()

def Excel():
    fileName = raw_input("请输入文件夹路径：")
    # 获取文件夹下所有文件
    _getAllFileOfPath(fileName)


def _getAllFileOfPath(fileName):
    ws = createExcel()
    index = 1
    for root, dirs, files in os.walk(fileName):
        # print("当前目录路径",root)  #当前目录路径
        # print("当前路径下的所有子目录",dirs)  # 当前路径下的所有子目录
        # print("当前路径下所有非目录子文件",files) # 当前路径下所有非目录子文件
        # 判断files 是否是excel
        for name in files:
            if name.endswith(".xls") or name.endswith(".xlsx"):
                print("是excel： %s ，文件路径为 %s" %(name,root))
                try:
                    restr = dealWithExcelFile(root + "/" + name,name)
                    index += 1
                    print(restr)
                    writeDataToExecl(ws, index, str(root[len(fileName):]),restr[0],restr[1] )
                except:
                    print("打开文件出错 excel： %s ，文件路径为 %s" %(name,root))
                # Useopenpyxl(root + "/" + name)

    wb.save("./抓取结果.xlsx")

def createExcel():
    ws = wb.active
    ws.insert_rows(1)
    ws["A1"] = "供应商"
    ws["B1"] = "文件名"
    ws["C1"] = "总金额"
    return ws


def writeDataToExecl(ws,index,supplier,fileName,totalPrice):
    ws.cell(index, 1).value = supplier
    ws.cell(index, 2).value = fileName
    ws.cell(index, 3).value = totalPrice



def dealWithExcelFile(excelFIle,fileName):
    # 打开文件
    data = xlrd.open_workbook(excelFIle)
    # 取出第一个表
    table = data.sheet_by_index(0)
    # 记录金额对应的列
    priceCol = 0

    totalPrice = "0"

    for rowNum in range(table.nrows):
        # 读取行的值
        rowValue = table.row_values(rowNum)
        for colNum in range(table.ncols):
            priceRow = str(rowValue[colNum])
            # print(priceRow)
            if (priceRow.startswith("金额") and (not priceRow.startswith("金额大写"))) or (priceRow.startswith("含税金额")):
                priceCol = colNum
            if priceRow.startswith("合计") or priceRow.startswith("总计") or priceRow.lower().startswith("total"):
                # print("金额---- ：",str(rowValue[priceCol]))
                if totalPrice == "0":
                    totalPrice = str(rowValue[priceCol])

    return (fileName,totalPrice)


def Useopenpyxl(fileName):
    # 加载文件
    wb = load_workbook(fileName)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    # 最大行数
    maxRow = ws.max_row
    # 最大列数
    maxColumn = ws.max_column
    # 记录金额对应的列
    priceCol = ""

    for rowNum in range(1,maxRow+1):
        #读取行的值
        rowValue = ws[rowNum]
        # print(rowValue)
        for tumple in rowValue:
            # print(tumple)
            priceRow = tumple.value
            print(priceRow)
            if str(priceRow).startswith('金额'):
                if len(priceCol) == 0:
                    priceCol = tumple.column_letter
            if (str(priceRow).startswith("合计") or str(priceRow).startswith("总计")):
                # print("金额 ---",rowValue[priceCol].value)
                priceCol += str(rowNum)
                break

    lastPrice = ws[priceCol]
    print("金额 ---" ,lastPrice.value)



if __name__ == "__main__":
    Excel()